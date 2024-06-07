import pandas as pd
from openpyxl import load_workbook

# Function to filter columns with 'Y' in the corresponding _Y column
def filter_columns_with_Y(df):
    columns_to_keep = []
    for col in df.columns:
        if col.endswith('_Y') and 'Y' in df[col].values:
            original_col = col[:-2]
            columns_to_keep.append(original_col)
    filtered_df = df[columns_to_keep]
    return filtered_df

# Function to adjust the column width
def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

def process_excel_file(input_file, output_file):
    try:
        # Read the original Excel file
        excel_data = pd.ExcelFile(input_file)
        
        # Process each sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(excel_data, sheet_name=sheet_name)
                filtered_df = filter_columns_with_Y(df)
                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Adjust column widths in the output file
        workbook = load_workbook(output_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            adjust_column_widths(sheet)
        workbook.save(output_file)
        
        print(f"Filtered data has been written to '{output_file}'")
    except Exception as e:
        print(f"An error occurred while processing the Excel file: {e}")

if __name__ == "__main__":
    try:
        input_file = 'purview_data.xlsx'
        output_file = 'filtered_purview_data.xlsx'
        process_excel_file(input_file, output_file)
    except Exception as e:
        print(f"An error occurred: {e}")
